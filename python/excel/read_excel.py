#!/usr/bin/env python  
# -*- coding:utf-8 -*-  
""" 
@file: read_excel.py
@time: 2023-02-14
@desc:
"""
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def get_data_from_excel(filepath):
    try:
        wb: Workbook = load_workbook(filepath)
        # 获取所有sheets
        sheets = wb.get_sheet_names()
        print(f"sheets: {sheets}")
        # 根据sheet名称获取sheet
        sheet = wb.get_sheet_by_name("Sheet1")
        print(f"sheet: {sheet}")
        # 获取第一个sheet
        ws: Worksheet = wb.active
        # 获取最大行数和列数
        max_row = ws.max_row
        max_column = ws.max_column
        print(f"max_row: {max_row}, max_column: {max_column}")
        # 获取数据 返回一个迭代器
        data = ws.values
        print(f"data: {data}")
        for row in ws.values:
            for value in row:
                print(value)
        values = list(data)
        print(f"values: {values}")
        return values
    except Exception as e:
        print(f"从excel读取数据失败,error: {str(e)}")


if __name__ == '__main__':
    get_data_from_excel("demo.xlsx")
