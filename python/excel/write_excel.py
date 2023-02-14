#!/usr/bin/env python  
# -*- coding:utf-8 -*-  
""" 
@file: write_excel.py
@time: 2023-02-14
@desc:
"""


def write_data_to_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    for row in data:
        ws.append(row)

    # 合并单元格
    ws.merge_cells("A1:D1")
    ws.merge_cells("F1:R1")
    alignment_center = Alignment(horizontal='center', vertical='center')
    # 指定区域单元格居中
    ws_area = ws["A1:R1"]
    font = Font(name="黑体", bold=True, size=18)
    for i in ws_area:
        for j in i:
            j.alignment = alignment_center
            j.font = font

    # # 指定区域单元格水平居中 垂直居中 自动换行
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_area = ws["A2:R2"]
    font = Font(name="黑体", bold=True, size=12)
    for i in ws_area:
        for j in i:
            j.alignment = alignment_center
            j.font = font

    # 设置列的宽度
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['F'].width = 15

    # Save the file
    wb.save("sample.xlsx")
    print("写入到Excel成功")
